from pathlib import Path
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path.cwd()
SOURCE = ROOT / "source.xlsx"
OUTDIR = ROOT / "outputs"
OUTPUT = OUTDIR / "daily_shipping_receivables_system.xlsx"


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def unique_preserve(values):
    seen = set()
    out = []
    for v in values:
        v = clean(v)
        if v is None:
            continue
        key = str(v)
        if key not in seen:
            seen.add(key)
            out.append(v)
    return out


def setup_sheet(ws, title, subtitle=None):
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color="1F2937")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Microsoft YaHei", size=10, color="6B7280")


def style_header(row):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_table_range(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Microsoft YaHei", size=10)


def add_table(ws, name, ref):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build():
    OUTDIR.mkdir(exist_ok=True)
    source_wb = load_workbook(SOURCE, data_only=True)
    air = source_wb["空运"]
    data_sheet = source_wb["数据获取和补充"]
    client_sheet = source_wb["客户信息"]

    channels = unique_preserve(air.cell(r, 2).value for r in range(3, air.max_row + 1))
    clients = unique_preserve(
        list(air.cell(r, 3).value for r in range(3, air.max_row + 1))
        + list(data_sheet.cell(r, 1).value for r in range(2, data_sheet.max_row + 1))
    )
    marks = unique_preserve(
        list(air.cell(r, 4).value for r in range(3, air.max_row + 1))
        + list(data_sheet.cell(r, 2).value for r in range(2, data_sheet.max_row + 1))
    )
    origins = unique_preserve(data_sheet.cell(r, 3).value for r in range(2, data_sheet.max_row + 1))
    destinations = unique_preserve(data_sheet.cell(r, 4).value for r in range(2, data_sheet.max_row + 1))
    terms = unique_preserve(data_sheet.cell(r, 5).value for r in range(2, data_sheet.max_row + 1))
    shipment_status = unique_preserve(data_sheet.cell(r, 6).value for r in range(2, data_sheet.max_row + 1))
    modes = unique_preserve(data_sheet.cell(r, 7).value for r in range(2, data_sheet.max_row + 1))
    payment_status = ["未付款N/P", "部分核销", "已付款PAID", "坏账/无需收款"]
    cost_status = ["未结算N/P", "已结算PAID", "部分结算", "无需结算"]
    currencies = ["RMB", "AED", "USD", "SAR"]

    wb = Workbook()
    ws_home = wb.active
    ws_home.title = "首页看板"
    ws_entry = wb.create_sheet("每日收货出货登记")
    ws_pay = wb.create_sheet("回款核销")
    ws_ar = wb.create_sheet("客户应收汇总")
    ws_open = wb.create_sheet("未核销明细")
    ws_client = wb.create_sheet("客户资料")
    ws_dict = wb.create_sheet("基础数据")

    setup_sheet(ws_home, "每日收货出货与应收回款看板", "登记每天收货/出货，回款按客户或运单号核销；统计页会自动更新。")
    setup_sheet(ws_entry, "每日收货出货登记")
    setup_sheet(ws_pay, "回款核销登记")
    setup_sheet(ws_ar, "客户应收汇总")
    setup_sheet(ws_open, "未核销明细")
    setup_sheet(ws_client, "客户资料")
    setup_sheet(ws_dict, "基础数据")

    entry_headers = [
        "录单日期", "业务类型", "渠道", "客户单位", "入仓号/唛头", "品名", "运单号", "始发地", "目的地", "运输方式",
        "件数", "实重", "体积重", "计费重量", "收款币种", "单价", "应收RMB", "结算方式", "货物状态", "派送日期",
        "付款状态", "已核销金额", "未收金额", "成本单价", "总计成本", "毛利润", "成本结算状态", "备注",
    ]
    ws_entry.append([""] * len(entry_headers))
    ws_entry.append(entry_headers)

    max_rows = 1000
    for r in range(3, max_rows + 3):
        ws_entry.cell(r, 17).value = f'=IF(OR(N{r}="",P{r}=""),"",N{r}*P{r})'
        ws_entry.cell(r, 22).value = f'=IF(G{r}="","",SUMIFS(回款核销!$H:$H,回款核销!$C:$C,G{r})+SUMIFS(回款核销!$H:$H,回款核销!$B:$B,D{r},回款核销!$C:$C,""))'
        ws_entry.cell(r, 23).value = f'=IF(Q{r}="","",MAX(0,Q{r}-V{r}))'
        ws_entry.cell(r, 25).value = f'=IF(OR(N{r}="",X{r}=""),"",N{r}*X{r})'
        ws_entry.cell(r, 26).value = f'=IF(OR(Q{r}="",Y{r}=""),"",Q{r}-Y{r})'
        ws_entry.cell(r, 21).value = f'=IF(Q{r}="","",IF(W{r}=0,"已付款PAID",IF(V{r}>0,"部分核销","未付款N/P")))'
    add_table(ws_entry, "tblEntry", f"A2:{get_column_letter(len(entry_headers))}{max_rows+2}")
    ws_entry.freeze_panes = "A3"
    style_header(ws_entry[2])
    style_table_range(ws_entry, 2, max_rows + 2, 1, len(entry_headers))
    set_widths(ws_entry, {
        "A": 12, "B": 12, "C": 12, "D": 18, "E": 18, "F": 22, "G": 16, "H": 12, "I": 12, "J": 16,
        "K": 8, "L": 10, "M": 10, "N": 10, "O": 10, "P": 10, "Q": 12, "R": 14, "S": 14, "T": 12,
        "U": 12, "V": 12, "W": 12, "X": 10, "Y": 12, "Z": 12, "AA": 14, "AB": 24,
    })

    pay_headers = ["回款日期", "客户单位", "运单号", "收款账户", "收款币种", "回款金额", "汇率", "核销金额RMB", "核销方式", "备注"]
    ws_pay.append([""] * len(pay_headers))
    ws_pay.append(pay_headers)
    for r in range(3, max_rows + 3):
        ws_pay.cell(r, 8).value = f'=IF(F{r}="","",F{r}*IF(G{r}="",1,G{r}))'
    add_table(ws_pay, "tblPayment", f"A2:J{max_rows+2}")
    ws_pay.freeze_panes = "A3"
    style_header(ws_pay[2])
    style_table_range(ws_pay, 2, max_rows + 2, 1, len(pay_headers))
    set_widths(ws_pay, {"A": 12, "B": 18, "C": 16, "D": 16, "E": 10, "F": 12, "G": 8, "H": 14, "I": 12, "J": 24})

    dict_cols = {
        "A": ("渠道", channels),
        "B": ("客户单位", clients),
        "C": ("货物唛头", marks),
        "D": ("始发地", origins),
        "E": ("目的地", destinations),
        "F": ("结算方式", terms),
        "G": ("货物状态", shipment_status),
        "H": ("运输方式", modes),
        "I": ("付款状态", payment_status),
        "J": ("成本结算状态", cost_status),
        "K": ("币种", currencies),
        "L": ("业务类型", ["收货", "出货", "退件", "费用调整"]),
        "M": ("核销方式", ["按运单核销", "按客户预收/批量核销", "退款/冲销", "其他"]),
    }
    for col, (header, vals) in dict_cols.items():
        ws_dict[f"{col}1"] = header
        ws_dict[f"{col}1"].font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        ws_dict[f"{col}1"].fill = PatternFill("solid", fgColor="1F4E79")
        for idx, val in enumerate(vals, start=2):
            ws_dict[f"{col}{idx}"] = val
        ws_dict.column_dimensions[col].width = 20
    ws_dict.freeze_panes = "A2"

    dv_specs = [
        ("B3:B1002", "基础数据!$L$2:$L$5"),
        ("C3:C1002", f"基础数据!$A$2:$A${max(2, len(channels)+1)}"),
        ("D3:D1002", f"基础数据!$B$2:$B${max(2, len(clients)+1)}"),
        ("E3:E1002", f"基础数据!$C$2:$C${max(2, len(marks)+1)}"),
        ("H3:H1002", f"基础数据!$D$2:$D${max(2, len(origins)+1)}"),
        ("I3:I1002", f"基础数据!$E$2:$E${max(2, len(destinations)+1)}"),
        ("J3:J1002", f"基础数据!$H$2:$H${max(2, len(modes)+1)}"),
        ("O3:O1002", "基础数据!$K$2:$K$5"),
        ("R3:R1002", f"基础数据!$F$2:$F${max(2, len(terms)+1)}"),
        ("S3:S1002", f"基础数据!$G$2:$G${max(2, len(shipment_status)+1)}"),
        ("U3:U1002", "基础数据!$I$2:$I$5"),
        ("AA3:AA1002", "基础数据!$J$2:$J$5"),
    ]
    for cell_range, formula in dv_specs:
        dv = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
        ws_entry.add_data_validation(dv)
        dv.add(cell_range)
    for cell_range, formula in [
        ("B3:B1002", f"基础数据!$B$2:$B${max(2, len(clients)+1)}"),
        ("E3:E1002", "基础数据!$K$2:$K$5"),
        ("I3:I1002", "基础数据!$M$2:$M$5"),
    ]:
        dv = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
        ws_pay.add_data_validation(dv)
        dv.add(cell_range)

    client_headers = ["唛头", "发货人", "发货人电话", "发货人地址", "收货人", "收货人电话", "收货人地址", "价格/备注"]
    ws_client.append([""] * len(client_headers))
    ws_client.append(client_headers)
    for row in client_sheet.iter_rows(min_row=3, values_only=True):
        if any(clean(v) is not None for v in row):
            ws_client.append(list(row[:8]))
    add_table(ws_client, "tblClients", f"A2:H{max(3, ws_client.max_row)}")
    style_header(ws_client[2])
    style_table_range(ws_client, 2, max(3, ws_client.max_row), 1, 8)
    set_widths(ws_client, {"A": 16, "B": 18, "C": 16, "D": 24, "E": 18, "F": 18, "G": 40, "H": 32})

    kpis = [
        ("本月应收", '=SUMIFS(每日收货出货登记!$Q:$Q,每日收货出货登记!$A:$A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),每日收货出货登记!$A:$A,"<"&EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),1))'),
        ("本月已核销", '=SUMIFS(每日收货出货登记!$V:$V,每日收货出货登记!$A:$A,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),每日收货出货登记!$A:$A,"<"&EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),1))'),
        ("当前未收", '=SUM(每日收货出货登记!$W:$W)'),
        ("当前毛利润", '=SUM(每日收货出货登记!$Z:$Z)'),
        ("未结算成本", '=SUMIFS(每日收货出货登记!$Y:$Y,每日收货出货登记!$AA:$AA,"<>已结算PAID")'),
    ]
    for idx, (label, formula) in enumerate(kpis):
        col = 1 + idx * 2
        ws_home.cell(4, col).value = label
        ws_home.cell(4, col).fill = PatternFill("solid", fgColor="EAF2F8")
        ws_home.cell(4, col).font = Font(name="Microsoft YaHei", bold=True, color="1F4E79")
        ws_home.cell(5, col).value = formula
        ws_home.cell(5, col).number_format = '#,##0.00'
        ws_home.cell(5, col).font = Font(name="Microsoft YaHei", size=14, bold=True, color="111827")
        ws_home.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        ws_home.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)

    ws_home["A8"] = "最近未收明细"
    ws_home["A8"].font = Font(name="Microsoft YaHei", size=13, bold=True)
    recent_headers = ["录单日期", "客户单位", "运单号", "应收RMB", "已核销", "未收", "货物状态"]
    for c, h in enumerate(recent_headers, 1):
        ws_home.cell(9, c).value = h
    style_header(ws_home[9])
    for r in range(10, 25):
        src = r - 7
        formulas = [
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!A{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!D{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!G{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!Q{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!V{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!W{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!S{src},"")',
        ]
        for c, f in enumerate(formulas, 1):
            ws_home.cell(r, c).value = f
    style_table_range(ws_home, 9, 24, 1, 7)
    set_widths(ws_home, {"A": 13, "B": 18, "C": 16, "D": 13, "E": 13, "F": 13, "G": 14, "I": 14, "J": 14})

    ws_home["I8"] = "使用顺序"
    ws_home["I8"].font = Font(name="Microsoft YaHei", size=13, bold=True)
    notes = [
        "1. 在“每日收货出货登记”新增每天业务。",
        "2. 收到款后在“回款核销”登记。",
        "3. 有运单号就按运单核销；批量收款可只填客户单位。",
        "4. 看“客户应收汇总”和“未核销明细”跟进欠款。",
    ]
    for i, note in enumerate(notes, 9):
        ws_home.cell(i, 9).value = note
        ws_home.cell(i, 9).font = Font(name="Microsoft YaHei", size=10, color="374151")
    ws_home.merge_cells("I9:J9")
    ws_home.merge_cells("I10:J10")
    ws_home.merge_cells("I11:J11")
    ws_home.merge_cells("I12:J12")

    ar_headers = ["客户单位", "总应收RMB", "已核销RMB", "未收RMB", "票数", "未收票数", "最后录单日期"]
    ws_ar.append([""] * len(ar_headers))
    ws_ar.append(ar_headers)
    for r in range(3, 303):
        ws_ar.cell(r, 1).value = f'=IFERROR(INDEX(基础数据!$B$2:$B$500,ROW(A{r-2})),"")'
        ws_ar.cell(r, 2).value = f'=IF(A{r}="","",SUMIFS(每日收货出货登记!$Q:$Q,每日收货出货登记!$D:$D,A{r}))'
        ws_ar.cell(r, 3).value = f'=IF(A{r}="","",SUMIFS(每日收货出货登记!$V:$V,每日收货出货登记!$D:$D,A{r}))'
        ws_ar.cell(r, 4).value = f'=IF(A{r}="","",SUMIFS(每日收货出货登记!$W:$W,每日收货出货登记!$D:$D,A{r}))'
        ws_ar.cell(r, 5).value = f'=IF(A{r}="","",COUNTIFS(每日收货出货登记!$D:$D,A{r},每日收货出货登记!$G:$G,"<>"))'
        ws_ar.cell(r, 6).value = f'=IF(A{r}="","",COUNTIFS(每日收货出货登记!$D:$D,A{r},每日收货出货登记!$W:$W,">0"))'
        ws_ar.cell(r, 7).value = f'=IF(A{r}="","",MAXIFS(每日收货出货登记!$A:$A,每日收货出货登记!$D:$D,A{r}))'
    add_table(ws_ar, "tblAR", "A2:G302")
    ws_ar.freeze_panes = "A3"
    style_header(ws_ar[2])
    style_table_range(ws_ar, 2, 302, 1, 7)
    set_widths(ws_ar, {"A": 22, "B": 14, "C": 14, "D": 14, "E": 10, "F": 10, "G": 14})

    open_headers = ["录单日期", "客户单位", "运单号", "入仓号/唛头", "应收RMB", "已核销", "未收", "结算方式", "货物状态", "备注"]
    ws_open.append([""] * len(open_headers))
    ws_open.append(open_headers)
    for r in range(3, 503):
        src = r
        formulas = [
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!A{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!D{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!G{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!E{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!Q{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!V{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!W{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!R{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!S{src},"")',
            f'=IF(每日收货出货登记!W{src}>0,每日收货出货登记!AB{src},"")',
        ]
        for c, f in enumerate(formulas, 1):
            ws_open.cell(r, c).value = f
    add_table(ws_open, "tblOpenAR", "A2:J502")
    ws_open.freeze_panes = "A3"
    style_header(ws_open[2])
    style_table_range(ws_open, 2, 502, 1, 10)
    set_widths(ws_open, {"A": 12, "B": 18, "C": 16, "D": 18, "E": 12, "F": 12, "G": 12, "H": 14, "I": 14, "J": 24})

    for ws in [ws_home, ws_entry, ws_pay, ws_ar, ws_open, ws_client]:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    if cell.column_letter in ["Q", "V", "W", "Y", "Z", "B", "C", "D", "E", "F", "H"]:
                        cell.number_format = '#,##0.00'
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    for ws in [ws_entry, ws_pay, ws_ar, ws_open]:
        ws.auto_filter.ref = ws.tables[next(iter(ws.tables))].ref

    for rng in ["Q3:Q1002", "V3:V1002", "W3:W1002", "Y3:Y1002", "Z3:Z1002"]:
        for row in ws_entry[rng]:
            for cell in row:
                cell.number_format = '#,##0.00'
    for rng in ["F3:H1002"]:
        for row in ws_pay[rng]:
            for cell in row:
                cell.number_format = '#,##0.00'
    for rng in ["B3:D302"]:
        for row in ws_ar[rng]:
            for cell in row:
                cell.number_format = '#,##0.00'
    for rng in ["E3:G502"]:
        for row in ws_open[rng]:
            for cell in row:
                cell.number_format = '#,##0.00'

    red_fill = PatternFill("solid", fgColor="FCE4D6")
    green_fill = PatternFill("solid", fgColor="E2F0D9")
    ws_entry.conditional_formatting.add("W3:W1002", CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill))
    ws_entry.conditional_formatting.add("U3:U1002", FormulaRule(formula=['$U3="已付款PAID"'], fill=green_fill))
    ws_ar.conditional_formatting.add("D3:D302", CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill))
    ws_open.conditional_formatting.add("G3:G502", CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill))

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "客户未收款排行"
    chart.y_axis.title = "客户"
    chart.x_axis.title = "未收RMB"
    chart.height = 7
    chart.width = 13
    data = Reference(ws_ar, min_col=4, min_row=2, max_row=12)
    cats = Reference(ws_ar, min_col=1, min_row=3, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_home.add_chart(chart, "I15")

    for ws in wb.worksheets:
        ws.freeze_panes = ws.freeze_panes
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 22
    ws_entry.row_dimensions[2].height = 42
    ws_pay.row_dimensions[2].height = 36
    ws_ar.row_dimensions[2].height = 34
    ws_open.row_dimensions[2].height = 34
    ws_client.row_dimensions[2].height = 34

    wb.save(OUTPUT)


if __name__ == "__main__":
    build()
